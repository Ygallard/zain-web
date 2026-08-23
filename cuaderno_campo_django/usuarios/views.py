import csv
from datetime import timedelta
from importlib.util import find_spec

from django.contrib import messages
from django.contrib.auth.hashers import check_password
from decimal import Decimal
from django.db.models import Count, DecimalField, ExpressionWrapper, F, FloatField, Q, Sum, Value
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.shortcuts import redirect, render
from django.utils import timezone

from .audit import IMPORTANT_AUDIT_ACTIONS, registrar_auditoria
from .decorators import AUTH_SESSION_KEY, access_denied_response, login_required_custom, role_required, any_authenticated
from .permissions import (
    get_filtered_predios,
    get_filtered_cuarteles,
    get_filtered_riegos,
    get_filtered_fertilizaciones,
    get_filtered_cosechas,
    get_filtered_aplicaciones_quimicas,
    user_owns_predio,
    user_owns_cuartel,
    user_owns_riego,
    user_owns_fertilizacion,
    user_owns_cosecha,
    user_owns_aplicacion_quimica,
    user_owns_labor_agricola,
    can_manage_predios,
    can_manage_cuarteles,
    can_manage_riegos,
    can_manage_fertilizaciones,
    can_manage_cosechas,
    can_manage_aplicaciones_quimicas,
    can_manage_labores_agricolas,
    get_filtered_labores_agricolas,
    get_sidebar_context,
    can_comment,
    can_send_notifications,
    get_registro_y_productor,
    user_can_view_registro,
)
from .forms import (
    AplicacionQuimicaCreateForm,
    AplicacionQuimicaUpdateForm,
    ComentarioTecnicoForm,
    CosechaCreateForm,
    CosechaUpdateForm,
    CuartelCreateForm,
    CuartelUpdateForm,
    FertilizacionCreateForm,
    FertilizacionUpdateForm,
    LoginForm,
    NotificacionForm,
    PredioCreateForm,
    PredioUpdateForm,
    RiegoCreateForm,
    RiegoUpdateForm,
    LaborAgricolaCreateForm,
    LaborAgricolaUpdateForm,
    UsuarioCreateForm,
    UsuarioUpdateForm,
)
from .models import (
    AplicacionQuimica,
    ComentarioTecnico,
    Cosecha,
    Cuartel,
    Fertilizacion,
    LaborAgricola,
    LogActividad,
    Notificacion,
    Predio,
    Riego,
    Usuario,
)


def get_current_user(request):
    user_id = request.session.get(AUTH_SESSION_KEY)
    if not user_id:
        return None
    return Usuario.objects.filter(id=user_id, estado=True).first()


def normalize_estado_post_data(request):
    data = request.POST.copy()
    if "estado" not in data:
        data["estado"] = "0"
    return data


def deny_if_cannot_manage(request, can_manage, message):
    if can_manage:
        return None
    return access_denied_response(request, message)


def get_dashboard_filters(request):
    return {
        "predio_id": request.GET.get("predio", "").strip(),
        "cuartel_id": request.GET.get("cuartel", "").strip(),
        "fecha_desde": request.GET.get("fecha_desde", "").strip(),
        "fecha_hasta": request.GET.get("fecha_hasta", "").strip(),
    }


def apply_common_filters(queryset, filters, date_field):
    qs = queryset
    predio_id = filters.get("predio_id")
    cuartel_id = filters.get("cuartel_id")
    fecha_desde = filters.get("fecha_desde")
    fecha_hasta = filters.get("fecha_hasta")

    if predio_id:
        qs = qs.filter(cuartel__predio_id=predio_id)
    if cuartel_id:
        qs = qs.filter(cuartel_id=cuartel_id)
    if fecha_desde:
        qs = qs.filter(**{f"{date_field}__gte": fecha_desde})
    if fecha_hasta:
        qs = qs.filter(**{f"{date_field}__lte": fecha_hasta})

    return qs


def build_month_series(queryset, date_field, aggregate_name, aggregate_expression, months=12):
    first_day_this_month = timezone.localdate().replace(day=1)
    start_month = (first_day_this_month - timedelta(days=(months - 1) * 31)).replace(day=1)

    qs = queryset.filter(**{f"{date_field}__gte": start_month})
    rows = (
        qs.annotate(periodo=TruncMonth(date_field))
        .values("periodo")
        .annotate(valor=aggregate_expression)
        .order_by("periodo")
    )

    values_by_month = {
        (item["periodo"].date() if hasattr(item["periodo"], "date") else item["periodo"]).replace(day=1): float(item["valor"] or 0)
        for item in rows
    }

    labels = []
    series = []
    cursor = start_month
    for _ in range(months):
        labels.append(cursor.strftime("%m/%Y"))
        series.append(round(values_by_month.get(cursor, 0.0), 2))
        next_month = (cursor.replace(day=28) + timedelta(days=4)).replace(day=1)
        cursor = next_month

    return {
        "name": aggregate_name,
        "labels": labels,
        "series": series,
    }


def login_view(request):
    if request.session.get(AUTH_SESSION_KEY):
        return redirect("dashboard")

    form = LoginForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        usuario_input = form.cleaned_data["usuario"].strip()
        password_input = form.cleaned_data["password"]

        usuario = Usuario.objects.filter(usuario=usuario_input, estado=True).first()

        if not usuario or not check_password(password_input, usuario.password):
            messages.error(request, "Credenciales inválidas.")
        else:
            request.session[AUTH_SESSION_KEY] = usuario.id
            request.session["rol"] = usuario.rol
            request.session["nombre"] = usuario.nombre
            request.session["usuario"] = usuario.usuario
            registrar_auditoria(
                usuario,
                LogActividad.ACCION_LOGIN,
                "autenticacion",
                f"{usuario.nombre} inicio sesion en la plataforma.",
            )
            messages.success(request, "Sesión iniciada correctamente.")
            return redirect("dashboard")

    return render(request, "login.html", {"form": form})


@login_required_custom
def logout_view(request):
    current_user = get_current_user(request)
    if current_user:
        registrar_auditoria(
            current_user,
            LogActividad.ACCION_LOGOUT,
            "autenticacion",
            f"{current_user.nombre} cerro sesion en la plataforma.",
        )

    request.session.flush()
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect("login")


@login_required_custom
def dashboard_view(request):
    current_user = get_current_user(request)
    if not current_user:
        messages.error(request, "Tu sesión no es válida. Vuelve a iniciar sesión.")
        return redirect("login")

    context = {
        "current_user": current_user,
        "sidebar_context": get_sidebar_context(request),
    }
    
    # Datos específicos por rol
    if current_user.rol == Usuario.ROL_ADMIN:
        ultima_actividad = (
            LogActividad.objects.select_related("usuario")
            .order_by("-fecha", "-id")
            .first()
        )

        context.update({
            "total_usuarios": Usuario.objects.count(),
            "usuarios_activos": Usuario.objects.filter(estado=True).count(),
            "usuarios_inactivos": Usuario.objects.filter(estado=False).count(),
            "administradores": Usuario.objects.filter(rol=Usuario.ROL_ADMIN).count(),
            "predios_count": Predio.objects.count(),
            "cuarteles_count": Cuartel.objects.filter(estado=True).count(),
            "total_riegos": Riego.objects.count(),
            "total_fertilizaciones": Fertilizacion.objects.count(),
            "total_cosechas": Cosecha.objects.count(),
            "total_kg_cosechados": Cosecha.objects.aggregate(total=Sum("cantidad_kg"))["total"],
            "total_bins_cosechados": Cosecha.objects.aggregate(total=Sum("cantidad_bins"))["total"],
            "total_aplicaciones_quimicas": AplicacionQuimica.objects.count(),
            "productos_quimicos_utilizados": AplicacionQuimica.objects.exclude(producto__isnull=True)
                .exclude(producto__exact="")
                .values("producto")
                .distinct()
                .count(),
            "cuarteles_tratados_quimicos": AplicacionQuimica.objects.values("cuartel_id").distinct().count(),
            "riegos_ultimos_30_dias": Riego.objects.filter(
                fecha_riego__gte=timezone.now().date() - timedelta(days=30)
            ).count(),
            "ultima_actividad_dashboard": (
                f"{ultima_actividad.usuario.nombre} - {ultima_actividad.modulo} - {ultima_actividad.accion}"
                if ultima_actividad
                else "-"
            ),
        })
        
        # Última cosecha y aplicación química
        ultima_cosecha_obj = Cosecha.objects.select_related("cuartel").order_by("-fecha_cosecha", "-created_at", "-id").first()
        context["ultima_cosecha_dashboard"] = (
            f"{ultima_cosecha_obj.cuartel.nombre_cuartel} - {ultima_cosecha_obj.fecha_cosecha.strftime('%d/%m/%Y')}"
            if ultima_cosecha_obj
            else "-"
        )
        
        ultima_aplicacion_obj = (
            AplicacionQuimica.objects.select_related("cuartel")
            .order_by("-fecha_aplicacion", "-created_at", "-id")
            .first()
        )
        context["ultima_aplicacion_quimica_dashboard"] = (
            f"{ultima_aplicacion_obj.cuartel.nombre_cuartel} - {ultima_aplicacion_obj.fecha_aplicacion.strftime('%d/%m/%Y')}"
            if ultima_aplicacion_obj
            else "-"
        )
    
    elif current_user.rol == Usuario.ROL_TECNICO:
        # PRODESAL: supervision y apoyo tecnico sobre todos los productores
        hoy = timezone.localdate()
        hace_7_dias = hoy - timedelta(days=6)

        context.update({
            "productores_supervisados": Usuario.objects.filter(rol=Usuario.ROL_PRODUCTOR, estado=True).count(),
            "predios_count": Predio.objects.count(),
            "cuarteles_count": Cuartel.objects.filter(estado=True).count(),
            "total_cosechas": Cosecha.objects.count(),
            "total_riegos": Riego.objects.count(),
            "total_fertilizaciones": Fertilizacion.objects.count(),
            "total_aplicaciones_quimicas": AplicacionQuimica.objects.count(),
            "riegos_recientes": Riego.objects.filter(fecha_riego__gte=hace_7_dias).count(),
            "fertilizaciones_recientes": Fertilizacion.objects.filter(fecha_aplicacion__gte=hace_7_dias).count(),
            "aplicaciones_recientes": AplicacionQuimica.objects.filter(fecha_aplicacion__gte=hace_7_dias).count(),
            "notificaciones_enviadas_pendientes": Notificacion.objects.filter(
                usuario_generador=current_user, leido=False
            ).count(),
            "observaciones_realizadas": ComentarioTecnico.objects.filter(usuario_prodesal=current_user).count(),
        })

    elif current_user.rol == Usuario.ROL_PRODUCTOR:
        # Productor solo ve sus datos
        mis_predios = Predio.objects.filter(usuario_id=current_user.id)
        mis_cuarteles = Cuartel.objects.filter(predio__usuario_id=current_user.id)
        mis_riegos = Riego.objects.filter(cuartel__predio__usuario_id=current_user.id)
        mis_fertilizaciones = Fertilizacion.objects.filter(cuartel__predio__usuario_id=current_user.id)
        mis_cosechas = Cosecha.objects.filter(cuartel__predio__usuario_id=current_user.id)
        mis_aplicaciones = AplicacionQuimica.objects.filter(cuartel__predio__usuario_id=current_user.id)
        
        context.update({
            "total_predios": mis_predios.count(),
            "total_cuarteles": mis_cuarteles.count(),
            "total_riegos": mis_riegos.count(),
            "total_fertilizaciones": mis_fertilizaciones.count(),
            "total_cosechas": mis_cosechas.count(),
            "total_kg_cosechados": mis_cosechas.aggregate(total=Sum("cantidad_kg"))["total"],
            "total_bins_cosechados": mis_cosechas.aggregate(total=Sum("cantidad_bins"))["total"],
            "total_aplicaciones_quimicas": mis_aplicaciones.count(),
            "notificaciones_no_leidas": Notificacion.objects.filter(productor=current_user, leido=False).count(),
            "observaciones_no_leidas": ComentarioTecnico.objects.filter(productor=current_user, leido=False).count(),
        })
    
    return render(request, "dashboard.html", context)


@any_authenticated
def dashboard_stats_api(request):
    current_user = get_current_user(request)
    if not current_user:
        return JsonResponse({"success": False, "error": "Sesion invalida"}, status=401)

    data = {
        "rol": current_user.rol,
        "future_alerts": [],
        "realtime_flowmeter_ready": True,
    }

    if current_user.rol == Usuario.ROL_ADMIN:
        ultima_actividad = (
            LogActividad.objects.select_related("usuario")
            .order_by("-fecha", "-id")
            .first()
        )
        data.update(
            {
                "usuarios_activos": Usuario.objects.filter(estado=True).count(),
                "predios_registrados": Predio.objects.count(),
                "cuarteles_activos": Cuartel.objects.filter(estado=True).count(),
                "total_riegos": Riego.objects.count(),
                "total_fertilizaciones": Fertilizacion.objects.count(),
                "total_cosechas": Cosecha.objects.count(),
                "total_aplicaciones_quimicas": AplicacionQuimica.objects.count(),
                "ultima_actividad": (
                    {
                        "usuario": ultima_actividad.usuario.nombre,
                        "modulo": ultima_actividad.modulo,
                        "accion": ultima_actividad.accion,
                        "fecha": ultima_actividad.fecha.strftime("%d/%m/%Y %H:%M") if ultima_actividad.fecha else "-",
                    }
                    if ultima_actividad
                    else None
                ),
            }
        )
    else:
        predios = get_filtered_predios(request)
        cuarteles = get_filtered_cuarteles(request)
        data.update(
            {
                "predios_registrados": predios.count(),
                "cuarteles_activos": cuarteles.filter(estado=True).count(),
                "total_riegos": get_filtered_riegos(request).count(),
                "total_fertilizaciones": get_filtered_fertilizaciones(request).count(),
                "total_cosechas": get_filtered_cosechas(request).count(),
                "total_aplicaciones_quimicas": get_filtered_aplicaciones_quimicas(request).count(),
            }
        )

    return JsonResponse({"success": True, "data": data})


@login_required_custom
def acceso_denegado_view(request):
    return render(
        request,
        "acceso_denegado.html",
        {
            "current_user": get_current_user(request),
            "sidebar_context": get_sidebar_context(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO)
def auditoria_logs_view(request):
    q = request.GET.get("q", "").strip()
    modulo = request.GET.get("modulo", "").strip()
    accion = request.GET.get("accion", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    logs = (
        LogActividad.objects.select_related("usuario")
        .filter(accion__in=IMPORTANT_AUDIT_ACTIONS)
        .order_by("-fecha", "-id")
    )

    if q:
        logs = logs.filter(
            Q(usuario__nombre__icontains=q)
            | Q(usuario__usuario__icontains=q)
            | Q(descripcion__icontains=q)
        )

    if modulo:
        logs = logs.filter(modulo__iexact=modulo)

    if accion:
        logs = logs.filter(accion__iexact=accion)

    if fecha_desde:
        logs = logs.filter(fecha__date__gte=fecha_desde)

    if fecha_hasta:
        logs = logs.filter(fecha__date__lte=fecha_hasta)

    paginator = Paginator(logs, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "current_user": get_current_user(request),
        "sidebar_context": get_sidebar_context(request),
        "page_obj": page_obj,
        "logs": page_obj.object_list,
        "q": q,
        "modulo": modulo,
        "accion": accion,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "modulos_disponibles": (
            LogActividad.objects.order_by("modulo")
            .values_list("modulo", flat=True)
            .distinct()
        ),
        "acciones_disponibles": sorted(IMPORTANT_AUDIT_ACTIONS),
        "total_logs": logs.count(),
    }
    return render(request, "auditoria_logs.html", context)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO)
def auditoria_log_detail_view(request, pk):
    try:
        log = LogActividad.objects.select_related("usuario").get(pk=pk)
        return JsonResponse(
            {
                "success": True,
                "log": {
                    "id": log.id,
                    "usuario_nombre": log.usuario.nombre,
                    "usuario_login": log.usuario.usuario,
                    "usuario_rol": log.usuario.rol,
                    "accion": log.accion,
                    "modulo": log.modulo,
                    "descripcion": log.descripcion or "",
                    "fecha": log.fecha.strftime("%d/%m/%Y") if log.fecha else "-",
                    "hora": log.fecha.strftime("%H:%M:%S") if log.fecha else "-",
                    "fecha_hora": log.fecha.strftime("%d/%m/%Y %H:%M:%S") if log.fecha else "-",
                },
            }
        )
    except LogActividad.DoesNotExist:
        return JsonResponse({"success": False, "error": "Registro de auditoría no encontrado"}, status=404)


@role_required(Usuario.ROL_ADMIN)
def usuario_list_view(request):
    current_user = get_current_user(request)

    q = request.GET.get("q", "").strip()
    queryset = Usuario.objects.all()

    if q:
        queryset = queryset.filter(
            Q(rut__icontains=q)
            | Q(nombre__icontains=q)
            | Q(usuario__icontains=q)
            | Q(rol__icontains=q)
            | Q(sector__icontains=q)
        )

    return render(
        request,
        "usuarios_lista.html",
        {
            "usuarios": queryset,
            "q": q,
            "current_user": current_user,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": True,
        },
    )


@role_required(Usuario.ROL_TECNICO)
def productores_lista_view(request):
    """Listado de solo lectura para que PRODESAL vea y notifique a los productores."""
    current_user = get_current_user(request)
    q = request.GET.get("q", "").strip()

    queryset = Usuario.objects.filter(rol=Usuario.ROL_PRODUCTOR)
    if q:
        queryset = queryset.filter(
            Q(rut__icontains=q)
            | Q(nombre__icontains=q)
            | Q(usuario__icontains=q)
            | Q(sector__icontains=q)
        )

    return render(
        request,
        "productores_lista.html",
        {
            "productores": queryset.order_by("nombre"),
            "q": q,
            "current_user": current_user,
            "sidebar_context": get_sidebar_context(request),
        },
    )


@role_required(Usuario.ROL_ADMIN)
def usuario_create_view(request):
    if request.method == "POST":
        try:
            form = UsuarioCreateForm(normalize_estado_post_data(request))
            if form.is_valid():
                usuario = form.save()
                return JsonResponse({
                    "success": True,
                    "message": "Usuario creado correctamente.",
                    "usuario": {
                        "id": usuario.id,
                        "rut": usuario.rut,
                        "nombre": usuario.nombre,
                        "usuario": usuario.usuario,
                        "rol": usuario.rol,
                        "celular": usuario.celular,
                        "sector": usuario.sector,
                        "estado": usuario.estado,
                    }
                })
            else:
                errors = {field: error[0] for field, error in form.errors.items()}
                return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    
    return redirect("usuarios_lista")


@role_required(Usuario.ROL_ADMIN)
def usuario_detail_view(request, pk):
    try:
        usuario = Usuario.objects.get(pk=pk)
        return JsonResponse({
            "success": True,
            "usuario": {
                "id": usuario.id,
                "rut": usuario.rut,
                "nombre": usuario.nombre,
                "usuario": usuario.usuario,
                "rol": usuario.rol,
                "celular": usuario.celular,
                "sector": usuario.sector,
                "estado": usuario.estado,
                "created_at": usuario.created_at.strftime("%d/%m/%Y %H:%M"),
            }
        })
    except Usuario.DoesNotExist:
        return JsonResponse({"success": False, "error": "Usuario no encontrado"}, status=404)


@role_required(Usuario.ROL_ADMIN)
def usuario_update_view(request, pk):
    if request.method == "POST":
        try:
            usuario = Usuario.objects.get(pk=pk)
            form = UsuarioUpdateForm(normalize_estado_post_data(request), instance=usuario)
            if form.is_valid():
                usuario = form.save()
                return JsonResponse({
                    "success": True,
                    "message": "Usuario actualizado correctamente.",
                    "usuario": {
                        "id": usuario.id,
                        "rut": usuario.rut,
                        "nombre": usuario.nombre,
                        "usuario": usuario.usuario,
                        "rol": usuario.rol,
                        "celular": usuario.celular,
                        "sector": usuario.sector,
                        "estado": usuario.estado,
                    }
                })
            else:
                errors = {field: error[0] for field, error in form.errors.items()}
                return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
        except Usuario.DoesNotExist:
            return JsonResponse({"success": False, "error": "Usuario no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    
    return redirect("usuarios_lista")


@role_required(Usuario.ROL_ADMIN)
def usuario_delete_view(request, pk):
    if request.method == "POST":
        try:
            usuario = Usuario.objects.get(pk=pk)
            nombre = usuario.nombre
            usuario.delete()
            return JsonResponse({
                "success": True,
                "message": f"Usuario {nombre} eliminado correctamente."
            })
        except Usuario.DoesNotExist:
            return JsonResponse({"success": False, "error": "Usuario no encontrado"}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)


@role_required(Usuario.ROL_ADMIN)
def usuario_toggle_estado_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    try:
        usuario = Usuario.objects.get(pk=pk)

        if usuario.id == request.session.get(AUTH_SESSION_KEY):
            return JsonResponse({"success": False, "error": "No puedes desactivar tu propia cuenta."}, status=400)

        usuario.estado = not usuario.estado
        usuario.save(update_fields=["estado"])

        estado_txt = "activado" if usuario.estado else "desactivado"
        return JsonResponse(
            {
                "success": True,
                "message": f"Usuario {usuario.nombre} {estado_txt} correctamente.",
                "estado": usuario.estado,
            }
        )
    except Usuario.DoesNotExist:
        return JsonResponse({"success": False, "error": "Usuario no encontrado"}, status=404)


ROL_LABELS = {
    Usuario.ROL_ADMIN: "Administrador",
    Usuario.ROL_TECNICO: "PRODESAL",
    Usuario.ROL_PRODUCTOR: "Productor",
}


def build_responsable_info(usuario):
    """Datos del usuario/productor que realmente es dueño del registro (no quien lo consulta)."""
    if not usuario:
        return {"responsable_id": None, "responsable_nombre": "-", "responsable_rol": "-"}

    return {
        "responsable_id": usuario.id,
        "responsable_nombre": usuario.nombre,
        "responsable_rol": ROL_LABELS.get(usuario.rol, usuario.rol),
    }


def build_ubicacion_info(predio, cuartel_nombre=""):
    """Ubicacion asociada a un registro a traves de su predio (nombre, sector, coordenadas)."""
    if not predio:
        return {
            "ubicacion_predio_nombre": "",
            "ubicacion_sector": "",
            "ubicacion_direccion": "",
            "ubicacion_cuartel": cuartel_nombre or "",
            "ubicacion_lat": "",
            "ubicacion_lng": "",
            "ubicacion_disponible": False,
        }

    lat = predio.geolocalizacion_lat
    lng = predio.geolocalizacion_lng

    return {
        "ubicacion_predio_nombre": predio.nombre_predio,
        "ubicacion_sector": predio.usuario.sector if predio.usuario_id and predio.usuario.sector else "",
        "ubicacion_direccion": predio.ubicacion or "",
        "ubicacion_cuartel": cuartel_nombre or "",
        "ubicacion_lat": str(lat) if lat is not None else "",
        "ubicacion_lng": str(lng) if lng is not None else "",
        "ubicacion_disponible": bool(predio.ubicacion or lat is not None or lng is not None or cuartel_nombre),
    }


def serialize_predio(predio):
    return {
        "id": predio.id,
        "usuario_id": predio.usuario_id,
        "usuario_nombre": predio.usuario.nombre,
        "usuario_label": f"{predio.usuario.nombre} ({predio.usuario.usuario})",
        "nombre_predio": predio.nombre_predio,
        "ubicacion": predio.ubicacion or "",
        "superficie": str(predio.superficie) if predio.superficie is not None else "",
        "superficie_hectareas": str(predio.superficie_hectareas) if predio.superficie_hectareas is not None else "",
        "inscripcion_cbr": predio.inscripcion_cbr or "",
        "inscripcion_agua": predio.inscripcion_agua or "",
        "geolocalizacion_lat": str(predio.geolocalizacion_lat) if predio.geolocalizacion_lat is not None else "",
        "geolocalizacion_lng": str(predio.geolocalizacion_lng) if predio.geolocalizacion_lng is not None else "",
        "descripcion": predio.descripcion or "",
        "estado": bool(predio.estado),
        "created_at": predio.created_at.strftime("%d/%m/%Y %H:%M") if predio.created_at else "-",
        **build_responsable_info(predio.usuario),
        **build_ubicacion_info(predio),
    }



@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def predio_list_view(request):
    current_user = get_current_user(request)
    predios = get_filtered_predios(request)
    can_manage = can_manage_predios(request)
    usuarios_activos = Usuario.objects.none()

    if can_manage:
        usuarios_activos = Usuario.objects.filter(estado=True).order_by("nombre")
        if current_user and current_user.rol == Usuario.ROL_PRODUCTOR:
            usuarios_activos = usuarios_activos.filter(id=current_user.id)

    return render(
        request,
        "predios_lista.html",
        {
            "predios": predios,
            "usuarios_activos": usuarios_activos,
            "current_user": current_user,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage,
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def predio_create_view(request):
    if request.method != "POST":
        return redirect("predios_lista")

    try:
        form = PredioCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            predio = form.save()
            predio.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Predio creado correctamente.",
                    "predio": serialize_predio(predio),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def predio_detail_view(request, pk):
    if not user_owns_predio(request, pk):
        return access_denied_response(request, "No autorizado para acceder a este predio.")

    try:
        predio = Predio.objects.select_related("usuario").get(pk=pk)
        return JsonResponse({"success": True, "predio": serialize_predio(predio)})
    except Predio.DoesNotExist:
        return JsonResponse({"success": False, "error": "Predio no encontrado"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def predio_update_view(request, pk):
    if request.method != "POST":
        return redirect("predios_lista")

    if not user_owns_predio(request, pk):
        return access_denied_response(request, "No autorizado para modificar este predio.")

    try:
        predio = Predio.objects.get(pk=pk)
        form = PredioUpdateForm(normalize_estado_post_data(request), instance=predio, request=request)
        if form.is_valid():
            predio = form.save()
            predio.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Predio actualizado correctamente.",
                    "predio": serialize_predio(predio),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Predio.DoesNotExist:
        return JsonResponse({"success": False, "error": "Predio no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def predio_delete_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_predio(request, pk):
        return access_denied_response(request, "No autorizado para eliminar este predio.")

    try:
        predio = Predio.objects.get(pk=pk)
        nombre = predio.nombre_predio
        predio.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Predio {nombre} eliminado correctamente.",
            }
        )
    except Predio.DoesNotExist:
        return JsonResponse({"success": False, "error": "Predio no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def predio_toggle_estado_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_predio(request, pk):
        return access_denied_response(request, "No autorizado para modificar este predio.")

    try:
        predio = Predio.objects.get(pk=pk)
        predio.estado = not bool(predio.estado)
        predio.save(update_fields=["estado"])
        estado_txt = "activado" if predio.estado else "desactivado"
        return JsonResponse(
            {
                "success": True,
                "message": f"Predio {predio.nombre_predio} {estado_txt} correctamente.",
                "estado": bool(predio.estado),
            }
        )
    except Predio.DoesNotExist:
        return JsonResponse({"success": False, "error": "Predio no encontrado"}, status=404)


# ---------------------------------------------------------------------------
# CUARTELES
# ---------------------------------------------------------------------------

def serialize_cuartel(cuartel):
    return {
        "id": cuartel.id,
        "predio_id": cuartel.predio_id,
        "predio_nombre": cuartel.predio.nombre_predio,
        "nombre_cuartel": cuartel.nombre_cuartel,
        "tipo_cultivo": cuartel.tipo_cultivo or "",
        "tipo_cultivo_label": cuartel.get_tipo_cultivo_display() if cuartel.tipo_cultivo else "",
        "variedad": cuartel.variedad or "",
        "forma_riego": cuartel.forma_riego or "",
        "forma_riego_label": cuartel.get_forma_riego_display() if cuartel.forma_riego else "-",
        "anio_plantacion": cuartel.anio_plantacion or "",
        "superficie": str(cuartel.superficie) if cuartel.superficie is not None else "",
        "descripcion": cuartel.descripcion or "",
        "estado": bool(cuartel.estado),
        "created_at": cuartel.created_at.strftime("%d/%m/%Y %H:%M") if cuartel.created_at else "-",
        **build_responsable_info(cuartel.predio.usuario),
        **build_ubicacion_info(cuartel.predio, cuartel.nombre_cuartel),
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cuartel_list_view(request):
    current_user = get_current_user(request)
    cuarteles = get_filtered_cuarteles(request)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    can_manage = can_manage_cuarteles(request)

    return render(
        request,
        "cuarteles_lista.html",
        {
            "cuarteles": cuarteles,
            "predios_activos": predios_activos,
            "current_user": current_user,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage,
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cuartel_create_view(request):
    if request.method != "POST":
        return redirect("cuarteles_lista")

    try:
        form = CuartelCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            cuartel = form.save()
            cuartel.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Cuartel creado correctamente.",
                    "cuartel": serialize_cuartel(cuartel),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cuartel_detail_view(request, pk):
    if not user_owns_cuartel(request, pk):
        return access_denied_response(request, "No autorizado para acceder a este cuartel.")

    try:
        cuartel = Cuartel.objects.select_related("predio", "predio__usuario").get(pk=pk)
        return JsonResponse({"success": True, "cuartel": serialize_cuartel(cuartel)})
    except Cuartel.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cuartel no encontrado"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cuartel_update_view(request, pk):
    if request.method != "POST":
        return redirect("cuarteles_lista")

    if not user_owns_cuartel(request, pk):
        return access_denied_response(request, "No autorizado para modificar este cuartel.")

    try:
        cuartel = Cuartel.objects.get(pk=pk)
        form = CuartelUpdateForm(normalize_estado_post_data(request), instance=cuartel, request=request)
        if form.is_valid():
            cuartel = form.save()
            cuartel.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Cuartel actualizado correctamente.",
                    "cuartel": serialize_cuartel(cuartel),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Cuartel.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cuartel no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cuartel_delete_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_cuartel(request, pk):
        return access_denied_response(request, "No autorizado para eliminar este cuartel.")

    try:
        cuartel = Cuartel.objects.get(pk=pk)
        nombre = cuartel.nombre_cuartel
        cuartel.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Cuartel {nombre} eliminado correctamente.",
            }
        )
    except Cuartel.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cuartel no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cuartel_toggle_estado_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_cuartel(request, pk):
        return access_denied_response(request, "No autorizado para cambiar estado de este cuartel.")

    try:
        cuartel = Cuartel.objects.get(pk=pk)
        cuartel.estado = not bool(cuartel.estado)
        cuartel.save(update_fields=["estado"])
        estado_txt = "activado" if cuartel.estado else "desactivado"
        return JsonResponse(
            {
                "success": True,
                "message": f"Cuartel {cuartel.nombre_cuartel} {estado_txt} correctamente.",
                "estado": bool(cuartel.estado),
            }
        )
    except Cuartel.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cuartel no encontrado"}, status=404)


# ---------------------------------------------------------------------------
# RIEGO
# ---------------------------------------------------------------------------

def build_riego_resumen(queryset=None):
    qs = queryset if queryset is not None else Riego.objects.all()
    litros_expr = ExpressionWrapper(
        (Coalesce(F("minutos_riego"), Value(0.0)) / Value(60.0)) * Coalesce(F("caudal"), Value(0.0)),
        output_field=FloatField(),
    )

    hoy = timezone.localdate()
    inicio_7_dias = hoy - timedelta(days=6)
    inicio_30_dias = hoy - timedelta(days=29)

    total_riegos = qs.count()
    total_litros = qs.aggregate(total=Sum(litros_expr))["total"] or 0
    riegos_ultimos_7_dias = qs.filter(fecha_riego__range=(inicio_7_dias, hoy)).count()
    riegos_ultimos_30_dias = qs.filter(fecha_riego__range=(inicio_30_dias, hoy)).count()

    cuartel_mas_regado = (
        qs.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total_litros=Sum(litros_expr), total_riegos=Count("id"))
        .order_by("-total_litros", "-total_riegos", "cuartel__nombre_cuartel")
        .first()
    )

    cuartel_mas_regado_label = "-"
    if cuartel_mas_regado:
        cuartel_mas_regado_label = (
            f"{cuartel_mas_regado['cuartel__nombre_cuartel']}"
            f" ({cuartel_mas_regado['cuartel__predio__nombre_predio']})"
        )

    ultimo_riego_obj = qs.select_related("cuartel", "cuartel__predio").order_by("-fecha_riego", "-created_at", "-id").first()
    ultimo_riego = (
        f"{ultimo_riego_obj.cuartel.nombre_cuartel} - {ultimo_riego_obj.fecha_riego.strftime('%d/%m/%Y')}"
        if ultimo_riego_obj
        else "-"
    )

    return {
        "total_riegos": total_riegos,
        "total_litros": round(float(total_litros), 2),
        "riegos_ultimos_7_dias": riegos_ultimos_7_dias,
        "riegos_ultimos_30_dias": riegos_ultimos_30_dias,
        "cuartel_mas_regado": cuartel_mas_regado_label,
        "ultimo_riego": ultimo_riego,
    }


def build_riego_analytics(queryset):
    litros_expr = ExpressionWrapper(
        (Coalesce(F("minutos_riego"), Value(0.0)) / Value(60.0)) * Coalesce(F("caudal"), Value(0.0)),
        output_field=FloatField(),
    )

    hoy = timezone.localdate()
    inicio_30_dias = hoy - timedelta(days=29)
    qs_30_dias = queryset.filter(fecha_riego__range=(inicio_30_dias, hoy))

    diario = {
        item["fecha_riego"]: {
            "total_riegos": item["total_riegos"],
            "total_litros": round(float(item["total_litros"] or 0), 2),
        }
        for item in (
            qs_30_dias.values("fecha_riego")
            .annotate(total_riegos=Count("id"), total_litros=Sum(litros_expr))
            .order_by("fecha_riego")
        )
    }

    labels_30 = []
    riegos_por_dia = []
    litros_por_dia = []
    for offset in range(30):
        fecha = inicio_30_dias + timedelta(days=offset)
        labels_30.append(fecha.strftime("%d/%m"))
        valores = diario.get(fecha, {"total_riegos": 0, "total_litros": 0.0})
        riegos_por_dia.append(valores["total_riegos"])
        litros_por_dia.append(valores["total_litros"])

    ranking_cuarteles = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total_litros=Sum(litros_expr), total_riegos=Count("id"))
        .order_by("-total_litros", "-total_riegos", "cuartel__nombre_cuartel")[:10]
    )

    ranking_labels = [
        f"{item['cuartel__nombre_cuartel']} ({item['cuartel__predio__nombre_predio']})"
        for item in ranking_cuarteles
    ]
    ranking_litros = [round(float(item["total_litros"] or 0), 2) for item in ranking_cuarteles]

    return {
        "riegos_por_dia": {
            "labels": labels_30,
            "series_riegos": riegos_por_dia,
            "series_litros": litros_por_dia,
        },
        "consumo_por_cuartel": {
            "labels": ranking_labels,
            "series_litros": ranking_litros,
        },
        "tendencia_30_dias": {
            "labels": labels_30,
            "series_litros": litros_por_dia,
        },
    }


def serialize_riego(riego):
    minutos = float(riego.minutos_riego or 0)
    caudal = float(riego.caudal or 0)
    litros = round((minutos / 60.0) * caudal, 2)

    return {
        "id": riego.id,
        "predio_id": riego.cuartel.predio_id,
        "predio_nombre": riego.cuartel.predio.nombre_predio,
        "responsable": riego.cuartel.predio.usuario.nombre,
        "cuartel_id": riego.cuartel_id,
        "cuartel_nombre": riego.cuartel.nombre_cuartel,
        "fecha_riego": riego.fecha_riego.strftime("%d/%m/%Y") if riego.fecha_riego else "-",
        "fecha_riego_iso": riego.fecha_riego.strftime("%Y-%m-%d") if riego.fecha_riego else "",
        "tipo_riego": riego.tipo_riego or "",
        "tipo_riego_label": riego.get_tipo_riego_display() if riego.tipo_riego else "",
        "minutos_riego": str(riego.minutos_riego) if riego.minutos_riego is not None else "",
        "caudal": str(riego.caudal) if riego.caudal is not None else "",
        "litros": litros,
        "observaciones": riego.observaciones or "",
        "estado": bool(riego.estado),
        "created_at": riego.created_at.strftime("%d/%m/%Y %H:%M") if riego.created_at else "-",
        **build_responsable_info(riego.cuartel.predio.usuario),
        **build_ubicacion_info(riego.cuartel.predio, riego.cuartel.nombre_cuartel),
    }





def get_riego_litros_expression():
    return ExpressionWrapper(
        (Coalesce(F("minutos_riego"), Value(0.0)) / Value(60.0)) * Coalesce(F("caudal"), Value(0.0)),
        output_field=FloatField(),
    )


def get_riego_gestion_queryset(request):
    q = request.GET.get("q", "").strip()
    predio_id = request.GET.get("predio", "").strip()
    cuartel_id = request.GET.get("cuartel", "").strip()
    estado = request.GET.get("estado", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    ordering = request.GET.get("ordering", "fecha_desc").strip()

    riegos = get_filtered_riegos(request)

    if q:
        riegos = riegos.filter(
            Q(cuartel__predio__nombre_predio__icontains=q)
            | Q(cuartel__nombre_cuartel__icontains=q)
            | Q(tipo_riego__icontains=q)
            | Q(observaciones__icontains=q)
            | Q(cuartel__predio__usuario__nombre__icontains=q)
        )

    if predio_id:
        riegos = riegos.filter(cuartel__predio_id=predio_id)

    if cuartel_id:
        riegos = riegos.filter(cuartel_id=cuartel_id)

    if estado in {"activo", "inactivo"}:
        riegos = riegos.filter(estado=(estado == "activo"))

    if fecha_desde:
        riegos = riegos.filter(fecha_riego__gte=fecha_desde)

    if fecha_hasta:
        riegos = riegos.filter(fecha_riego__lte=fecha_hasta)

    riegos = riegos.annotate(litros_estimados=get_riego_litros_expression())

    ordering_map = {
        "fecha_desc": ["-fecha_riego", "-id"],
        "fecha_asc": ["fecha_riego", "id"],
        "litros_desc": ["-litros_estimados", "-fecha_riego", "-id"],
        "litros_asc": ["litros_estimados", "fecha_riego", "id"],
        "cuartel_asc": ["cuartel__nombre_cuartel", "-fecha_riego", "-id"],
        "predio_asc": ["cuartel__predio__nombre_predio", "-fecha_riego", "-id"],
    }
    riegos = riegos.order_by(*ordering_map.get(ordering, ordering_map["fecha_desc"]))

    return riegos, {
        "q": q,
        "predio_id": predio_id,
        "cuartel_id": cuartel_id,
        "estado": estado,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "ordering": ordering,
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def riego_list_view(request):
    return redirect("riegos_gestion")


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def riego_gestion_view(request):
    current_user = get_current_user(request)
    riegos, filtros = get_riego_gestion_queryset(request)

    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True)
    if filtros["predio_id"]:
        cuarteles_activos = cuarteles_activos.filter(predio_id=filtros["predio_id"])
    cuarteles_activos = cuarteles_activos.order_by("nombre_cuartel")

    paginator = Paginator(riegos, 15)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "riegos_gestion.html",
        {
            "riegos": page_obj.object_list,
            "page_obj": page_obj,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_riegos(request),
            **filtros,
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def riego_analitica_view(request):
    current_user = get_current_user(request)
    riegos = get_filtered_riegos(request).annotate(litros_estimados=get_riego_litros_expression())
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")

    return render(
        request,
        "riegos_analitica.html",
        {
            "current_user": current_user,
            "resumen": build_riego_resumen(riegos),
            "analytics": build_riego_analytics(riegos),
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "sidebar_context": get_sidebar_context(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def riego_analytics_api(request):
    filters = get_dashboard_filters(request)
    qs = get_filtered_riegos(request).annotate(litros_estimados=get_riego_litros_expression())
    qs = apply_common_filters(qs, filters, "fecha_riego")

    return JsonResponse(
        {
            "success": True,
            "resumen": build_riego_resumen(qs),
            "analytics": build_riego_analytics(qs),
        }
    )


def can_export_riego_reports(request):
    rol = request.session.get("rol")
    return rol in {Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR}


def build_export_dependency_error_response(request, missing_packages):
    missing_str = ", ".join(missing_packages)
    message = (
        "La exportación no está disponible porque faltan dependencias: "
        f"{missing_str}. Ejecuta: pip install reportlab openpyxl pandas matplotlib"
    )

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"success": False, "error": message}, status=503)

    messages.error(request, message)
    return redirect("riegos_gestion")


def get_riego_export_dataset(request):
    riegos, filtros = get_riego_gestion_queryset(request)

    predio_label = "Todos"
    if filtros["predio_id"]:
        predio_obj = get_filtered_predios(request).filter(id=filtros["predio_id"]).first()
        predio_label = predio_obj.nombre_predio if predio_obj else "N/A"

    cuartel_label = "Todos"
    if filtros["cuartel_id"]:
        cuartel_obj = get_filtered_cuarteles(request).filter(id=filtros["cuartel_id"]).first()
        cuartel_label = cuartel_obj.nombre_cuartel if cuartel_obj else "N/A"

    estado_map = {
        "activo": "Activo",
        "inactivo": "Inactivo",
        "": "Todos",
    }

    filtros_aplicados = [
        ("Busqueda", filtros["q"] or "-"),
        ("Predio", predio_label),
        ("Cuartel", cuartel_label),
        ("Estado", estado_map.get(filtros["estado"], "Todos")),
        ("Fecha desde", filtros["fecha_desde"] or "-"),
        ("Fecha hasta", filtros["fecha_hasta"] or "-"),
    ]

    total_litros = riegos.aggregate(total=Sum("litros_estimados"))["total"] or 0

    return {
        "riegos": riegos,
        "filtros": filtros,
        "filtros_aplicados": filtros_aplicados,
        "total_registros": riegos.count(),
        "total_litros": round(float(total_litros), 2),
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_export_excel_view(request):
    if not can_export_riego_reports(request):
        return access_denied_response(request, "No tienes permisos para exportar reportes de riego.")

    if find_spec("openpyxl") is None:
        return build_export_dependency_error_response(request, ["openpyxl"])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception:
        return build_export_dependency_error_response(request, ["openpyxl"])

    dataset = get_riego_export_dataset(request)
    current_user = get_current_user(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Riego"

    ws["A1"] = "Cuaderno de Campo Digital"
    ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Reporte de Riego - Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells("A2:F2")
    ws["A3"] = f"Usuario: {current_user.nombre if current_user else '-'}"
    ws.merge_cells("A3:G3")

    ws["A5"] = "Filtros aplicados"
    ws["A5"].font = Font(bold=True)

    filtro_row = 6
    for label, value in dataset["filtros_aplicados"]:
        ws[f"A{filtro_row}"] = label
        ws[f"A{filtro_row}"].font = Font(bold=True)
        ws[f"B{filtro_row}"] = value
        ws.merge_cells(start_row=filtro_row, start_column=2, end_row=filtro_row, end_column=7)
        filtro_row += 1

    table_row = filtro_row + 1
    headers = ["Fecha", "Predio", "Cuartel", "Minutos", "Litros", "Responsable", "Observaciones"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_row, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    data_row = table_row + 1
    for riego in dataset["riegos"]:
        litros = float(getattr(riego, "litros_estimados", 0) or 0)
        ws.cell(data_row, 1, riego.fecha_riego.strftime("%d/%m/%Y") if riego.fecha_riego else "-")
        ws.cell(data_row, 2, riego.cuartel.predio.nombre_predio)
        ws.cell(data_row, 3, riego.cuartel.nombre_cuartel)
        ws.cell(data_row, 4, float(riego.minutos_riego or 0))
        ws.cell(data_row, 5, litros)
        ws.cell(data_row, 6, riego.cuartel.predio.usuario.nombre)
        ws.cell(data_row, 7, riego.observaciones or "-")

        ws.cell(data_row, 4).number_format = "#,##0.00"
        ws.cell(data_row, 5).number_format = "#,##0.00"
        for col in range(1, 8):
            ws.cell(data_row, col).border = thin_border
            ws.cell(data_row, col).alignment = Alignment(vertical="top", wrap_text=True)
        data_row += 1

    summary_row = data_row + 1
    ws.cell(summary_row, 1, "Total registros")
    ws.cell(summary_row, 1).font = Font(bold=True)
    ws.cell(summary_row, 2, dataset["total_registros"])

    ws.cell(summary_row + 1, 1, "Litros totales")
    ws.cell(summary_row + 1, 1).font = Font(bold=True)
    ws.cell(summary_row + 1, 2, dataset["total_litros"])
    ws.cell(summary_row + 1, 2).number_format = "#,##0.00"

    column_widths = {
        "A": 14,
        "B": 24,
        "C": 24,
        "D": 14,
        "E": 14,
        "F": 24,
        "G": 42,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    filename = f"riego_reporte_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_export_pdf_view(request):
    if not can_export_riego_reports(request):
        return access_denied_response(request, "No tienes permisos para exportar reportes de riego.")

    if find_spec("reportlab") is None:
        return build_export_dependency_error_response(request, ["reportlab"])

    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        from reportlab.graphics.shapes import Drawing, Line, Rect, String
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        return build_export_dependency_error_response(request, ["reportlab"])

    dataset = get_riego_export_dataset(request)
    analytics = build_riego_analytics(dataset["riegos"])
    current_user = get_current_user(request)
    generated_at = timezone.localtime()
    generated_at_text = generated_at.strftime("%d/%m/%Y %H:%M")

    promedio_consumo = 0.0
    if dataset["total_registros"]:
        promedio_consumo = dataset["total_litros"] / dataset["total_registros"]

    filename = f"riego_reporte_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.pdf"

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=30 * mm,
        bottomMargin=24 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitleCustom",
        parent=styles["Heading1"],
        fontSize=15,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=8,
        leading=18,
    )
    subtitle_style = ParagraphStyle(
        "PdfSubtitle",
        parent=styles["Heading3"],
        fontSize=11,
        textColor=colors.HexColor("#2C3E50"),
        spaceAfter=6,
    )
    info_style = ParagraphStyle("PdfInfo", parent=styles["Normal"], fontSize=9, leading=12)
    body_style = ParagraphStyle("PdfBody", parent=styles["Normal"], fontSize=9.2, leading=13)

    def _truncate_labels(labels, max_len=24):
        return [f"{label[:max_len-1]}..." if len(label) > max_len else label for label in labels]

    def build_vertical_bar_chart(labels, values, title):
        chart_width = 180 * mm
        chart_height = 55 * mm
        d = Drawing(chart_width, chart_height)

        d.add(String(0, chart_height - 8, title, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#1F4E78")))
        if not labels:
            d.add(String(0, chart_height / 2, "Sin datos para el periodo seleccionado", fontSize=9, fillColor=colors.HexColor("#7F8C8D")))
            return d

        chart = VerticalBarChart()
        chart.x = 0
        chart.y = 0
        chart.height = 40 * mm
        chart.width = chart_width
        chart.data = [values]
        chart.categoryAxis.categoryNames = _truncate_labels(labels, max_len=8)
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.dy = -10
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        max_value = max(values) if values else 0
        chart.valueAxis.valueMax = max(max_value * 1.2, 10)
        chart.valueAxis.valueStep = max(chart.valueAxis.valueMax / 5, 1)
        chart.valueAxis.labels.fontSize = 7
        chart.barWidth = max((chart_width / max(len(values), 1)) * 0.55, 3)
        chart.groupSpacing = 2
        chart.bars[0].fillColor = colors.HexColor("#3498DB")
        chart.bars[0].strokeColor = colors.HexColor("#2C7FB8")
        d.add(chart)
        return d

    def build_horizontal_bar_chart(labels, values, title):
        chart_width = 180 * mm
        chart_height = 60 * mm
        d = Drawing(chart_width, chart_height)

        d.add(String(0, chart_height - 8, title, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#1F4E78")))
        if not labels:
            d.add(String(0, chart_height / 2, "Sin datos para el periodo seleccionado", fontSize=9, fillColor=colors.HexColor("#7F8C8D")))
            return d

        chart = HorizontalBarChart()
        chart.x = 45 * mm
        chart.y = 0
        chart.height = 42 * mm
        chart.width = 130 * mm
        chart.data = [values]
        chart.categoryAxis.categoryNames = _truncate_labels(labels, max_len=16)
        chart.categoryAxis.labels.fontSize = 7
        chart.valueAxis.valueMin = 0
        max_value = max(values) if values else 0
        chart.valueAxis.valueMax = max(max_value * 1.2, 10)
        chart.valueAxis.valueStep = max(chart.valueAxis.valueMax / 5, 1)
        chart.valueAxis.labels.fontSize = 7
        chart.bars[0].fillColor = colors.HexColor("#16A085")
        chart.bars[0].strokeColor = colors.HexColor("#0E7A66")
        d.add(chart)
        return d

    def build_line_chart(labels, values, title):
        chart_width = 180 * mm
        chart_height = 56 * mm
        d = Drawing(chart_width, chart_height)

        d.add(String(0, chart_height - 8, title, fontName="Helvetica-Bold", fontSize=10, fillColor=colors.HexColor("#1F4E78")))
        if not labels:
            d.add(String(0, chart_height / 2, "Sin datos para el periodo seleccionado", fontSize=9, fillColor=colors.HexColor("#7F8C8D")))
            return d

        chart = HorizontalLineChart()
        chart.x = 0
        chart.y = 0
        chart.height = 40 * mm
        chart.width = chart_width
        chart.data = [values]
        chart.categoryAxis.categoryNames = _truncate_labels(labels, max_len=8)
        chart.categoryAxis.labels.angle = 35
        chart.categoryAxis.labels.dy = -10
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        max_value = max(values) if values else 0
        chart.valueAxis.valueMax = max(max_value * 1.2, 10)
        chart.valueAxis.valueStep = max(chart.valueAxis.valueMax / 5, 1)
        chart.valueAxis.labels.fontSize = 7
        chart.lines[0].strokeColor = colors.HexColor("#F39C12")
        chart.lines[0].strokeWidth = 1.8
        chart.lines[0].symbol = None
        d.add(chart)
        return d

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_header_footer(total_pages)
                super().showPage()
            super().save()

        def draw_header_footer(self, total_pages):
            page_number = self._pageNumber
            width, height = A4

            # Header fixed
            self.saveState()
            self.setFillColor(colors.HexColor("#1F4E78"))
            self.rect(10 * mm, height - 20 * mm, width - 20 * mm, 10 * mm, fill=1, stroke=0)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 8.6)
            self.drawString(14 * mm, height - 14.8 * mm, "Zaino Web - Sistema de Monitoreo Agricola")
            self.setFont("Helvetica", 7.2)
            self.drawRightString(width - 14 * mm, height - 14.8 * mm, f"Generado: {generated_at_text} | Usuario: {current_user.nombre if current_user else '-'}")

            # pseudo-logo
            self.setFillColor(colors.HexColor("#3498DB"))
            self.circle(14 * mm, height - 15 * mm, 2.8 * mm, stroke=0, fill=1)
            self.setFillColor(colors.white)
            self.setFont("Helvetica-Bold", 5.5)
            self.drawCentredString(14 * mm, height - 16 * mm, "ZW")
            self.restoreState()

            # Footer fixed
            self.saveState()
            self.setStrokeColor(colors.HexColor("#D9E2EC"))
            self.setLineWidth(0.4)
            self.line(10 * mm, 17 * mm, width - 10 * mm, 17 * mm)

            self.setFillColor(colors.HexColor("#2C3E50"))
            self.setFont("Helvetica", 7.2)
            self.drawString(10 * mm, 12.2 * mm, "Zaino Web - Sistema de Monitoreo Agricola")
            self.drawString(10 * mm, 9.1 * mm, "Instituto Profesional AIEP")
            self.drawRightString(width - 10 * mm, 12.2 * mm, f"Pagina {page_number} de {total_pages}")
            self.drawRightString(width - 10 * mm, 9.1 * mm, "Documento generado automaticamente")
            self.restoreState()

    story = []
    story.append(Paragraph("Cuaderno de Campo Digital", title_style))
    story.append(Paragraph("Reporte Ejecutivo de Riego", subtitle_style))
    story.append(
        Paragraph(
            f"Generado: {generated_at_text} | "
            f"Usuario: {current_user.nombre if current_user else '-'}",
            info_style,
        )
    )
    story.append(Spacer(1, 4 * mm))

    summary_cards = [
        [
            Paragraph("<b>Total registros</b>", info_style),
            Paragraph("<b>Litros totales</b>", info_style),
            Paragraph("<b>Promedio consumo</b>", info_style),
            Paragraph("<b>Fecha exportacion</b>", info_style),
        ],
        [
            Paragraph(str(dataset["total_registros"]), styles["Heading3"]),
            Paragraph(f"{dataset['total_litros']:.2f} L", styles["Heading3"]),
            Paragraph(f"{promedio_consumo:.2f} L/registro", styles["Heading3"]),
            Paragraph(generated_at_text, styles["Heading3"]),
        ],
    ]
    cards_table = Table(summary_cards, colWidths=[44 * mm, 44 * mm, 44 * mm, 44 * mm])
    cards_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F7FC")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2EC")),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9E2EC")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(cards_table)
    story.append(Spacer(1, 4 * mm))

    filtros_text = "<br/>".join(
        f"<b>{label}:</b> {value}" for label, value in dataset["filtros_aplicados"]
    )
    story.append(Paragraph("<b>Resumen Ejecutivo</b>", subtitle_style))
    story.append(
        Paragraph(
            f"Durante el periodo seleccionado se registro un consumo total de agua de "
            f"<b>{dataset['total_litros']:.2f} litros</b> en "
            f"<b>{dataset['total_registros']}</b> registros de riego.",
            body_style,
        )
    )
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph("<b>Filtros aplicados</b>", info_style))
    story.append(Paragraph(filtros_text, info_style))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Visualizacion de Consumo</b>", subtitle_style))
    story.append(build_vertical_bar_chart(
        analytics["riegos_por_dia"]["labels"],
        analytics["riegos_por_dia"]["series_litros"],
        "Consumo diario de agua (litros)",
    ))
    story.append(Spacer(1, 3 * mm))
    story.append(build_line_chart(
        analytics["tendencia_30_dias"]["labels"],
        analytics["tendencia_30_dias"]["series_litros"],
        "Tendencia de riego ultimos 30 dias",
    ))
    story.append(Spacer(1, 3 * mm))
    story.append(build_horizontal_bar_chart(
        analytics["consumo_por_cuartel"]["labels"],
        analytics["consumo_por_cuartel"]["series_litros"],
        "Litros por cuartel",
    ))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("<b>Detalle de Registros</b>", subtitle_style))

    table_data = [["Fecha", "Predio", "Cuartel", "Minutos", "Litros", "Responsable", "Observaciones"]]
    for riego in dataset["riegos"]:
        table_data.append(
            [
                riego.fecha_riego.strftime("%d/%m/%Y") if riego.fecha_riego else "-",
                riego.cuartel.predio.nombre_predio,
                riego.cuartel.nombre_cuartel,
                f"{float(riego.minutos_riego or 0):.2f}",
                f"{float(getattr(riego, 'litros_estimados', 0) or 0):.2f}",
                riego.cuartel.predio.usuario.nombre,
                (riego.observaciones or "-")[:120],
            ]
        )

    col_widths = [20 * mm, 30 * mm, 26 * mm, 16 * mm, 16 * mm, 26 * mm, 46 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 4 * mm))
    story.append(
        Paragraph(
            f"<b>Total registros:</b> {dataset['total_registros']} &nbsp;&nbsp;&nbsp; "
            f"<b>Litros totales:</b> {dataset['total_litros']:.2f}",
            body_style,
        )
    )

    doc.build(story, canvasmaker=NumberedCanvas)
    return response


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cuarteles_por_predio_view(request, predio_id):
    if not get_filtered_predios(request).filter(id=predio_id, estado=True).exists():
        return access_denied_response(request, "No autorizado para consultar cuarteles de este predio.")

    cuarteles = (
        get_filtered_cuarteles(request).filter(predio_id=predio_id, predio__estado=True)
        .order_by("nombre_cuartel")
        .values("id", "nombre_cuartel")
    )
    return JsonResponse({"success": True, "cuarteles": list(cuarteles)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_create_view(request):
    if request.method != "POST":
        return redirect("riegos_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_riegos(request),
        "No tienes permisos para crear riegos.",
    )
    if denied_response:
        return denied_response

    try:
        form = RiegoCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            riego = form.save()
            riego.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Riego creado correctamente.",
                    "riego": serialize_riego(riego),
                    "resumen": build_riego_resumen(get_filtered_riegos(request)),
                    "analytics": build_riego_analytics(get_filtered_riegos(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def riego_detail_view(request, pk):
    if not user_owns_riego(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    try:
        riego = Riego.objects.select_related("cuartel", "cuartel__predio", "cuartel__predio__usuario").get(pk=pk)
        return JsonResponse({"success": True, "riego": serialize_riego(riego)})
    except Riego.DoesNotExist:
        return JsonResponse({"success": False, "error": "Riego no encontrado"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_update_view(request, pk):
    if not user_owns_riego(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return redirect("riegos_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_riegos(request),
        "No tienes permisos para editar riegos.",
    )
    if denied_response:
        return denied_response

    try:
        riego = Riego.objects.get(pk=pk)
        form = RiegoUpdateForm(normalize_estado_post_data(request), instance=riego, request=request)
        if form.is_valid():
            riego = form.save()
            riego.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Riego actualizado correctamente.",
                    "riego": serialize_riego(riego),
                    "resumen": build_riego_resumen(get_filtered_riegos(request)),
                    "analytics": build_riego_analytics(get_filtered_riegos(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Riego.DoesNotExist:
        return JsonResponse({"success": False, "error": "Riego no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_delete_view(request, pk):
    if not user_owns_riego(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_riegos(request),
        "No tienes permisos para eliminar riegos.",
    )
    if denied_response:
        return denied_response

    try:
        riego = Riego.objects.select_related("cuartel").get(pk=pk)
        nombre_cuartel = riego.cuartel.nombre_cuartel
        riego.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Riego del cuartel {nombre_cuartel} eliminado correctamente.",
                "resumen": build_riego_resumen(get_filtered_riegos(request)),
                "analytics": build_riego_analytics(get_filtered_riegos(request)),
            }
        )
    except Riego.DoesNotExist:
        return JsonResponse({"success": False, "error": "Riego no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def riego_toggle_estado_view(request, pk):
    if not user_owns_riego(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_riegos(request),
        "No tienes permisos para cambiar estado de riegos.",
    )
    if denied_response:
        return denied_response

    try:
        riego = Riego.objects.get(pk=pk)
        riego.estado = not bool(riego.estado)
        riego.save(update_fields=["estado"])
        estado_txt = "activado" if riego.estado else "desactivado"
        return JsonResponse(
            {
                "success": True,
                "message": f"Riego {riego.id} {estado_txt} correctamente.",
                "estado": bool(riego.estado),
                "resumen": build_riego_resumen(get_filtered_riegos(request)),
                "analytics": build_riego_analytics(get_filtered_riegos(request)),
            }
        )
    except Riego.DoesNotExist:
        return JsonResponse({"success": False, "error": "Riego no encontrado"}, status=404)


# ---------------------------------------------------------------------------
# FERTILIZACION
# ---------------------------------------------------------------------------

def build_fertilizacion_resumen(queryset=None):
    qs = queryset if queryset is not None else Fertilizacion.objects.all()
    total_aplicaciones = qs.count()
    productos_utilizados = qs.exclude(producto__isnull=True).exclude(producto__exact="").values("producto").distinct().count()
    cuarteles_fertilizados = qs.values("cuartel_id").distinct().count()
    ultima_aplicacion_obj = qs.select_related("cuartel").order_by("-fecha_aplicacion", "-created_at", "-id").first()
    ultima_aplicacion = (
        f"{ultima_aplicacion_obj.cuartel.nombre_cuartel} - {ultima_aplicacion_obj.fecha_aplicacion.strftime('%d/%m/%Y')}"
        if ultima_aplicacion_obj
        else "-"
    )

    return {
        "total_aplicaciones": total_aplicaciones,
        "productos_utilizados": productos_utilizados,
        "cuarteles_fertilizados": cuarteles_fertilizados,
        "ultima_aplicacion": ultima_aplicacion,
    }


def serialize_fertilizacion(fertilizacion):
    return {
        "id": fertilizacion.id,
        "predio_id": fertilizacion.cuartel.predio_id,
        "predio_nombre": fertilizacion.cuartel.predio.nombre_predio,
        "cuartel_id": fertilizacion.cuartel_id,
        "cuartel_nombre": fertilizacion.cuartel.nombre_cuartel,
        "fecha_aplicacion": fertilizacion.fecha_aplicacion.strftime("%d/%m/%Y") if fertilizacion.fecha_aplicacion else "-",
        "fecha_aplicacion_iso": fertilizacion.fecha_aplicacion.strftime("%Y-%m-%d") if fertilizacion.fecha_aplicacion else "",
        "producto": fertilizacion.producto or "",
        "producto_label": fertilizacion.get_producto_display() if fertilizacion.producto else "",
        "dosis": str(fertilizacion.dosis) if fertilizacion.dosis is not None else "",
        "unidad": fertilizacion.unidad or "",
        "metodo_aplicacion": fertilizacion.metodo_aplicacion or "",
        "observaciones": fertilizacion.observaciones or "",
        "estado": bool(fertilizacion.estado),
        "created_at": fertilizacion.created_at.strftime("%d/%m/%Y %H:%M") if fertilizacion.created_at else "-",
        **build_responsable_info(fertilizacion.cuartel.predio.usuario),
        **build_ubicacion_info(fertilizacion.cuartel.predio, fertilizacion.cuartel.nombre_cuartel),
    }


def build_fertilizacion_analytics(queryset):
    productos_top = list(
        queryset.exclude(producto__isnull=True)
        .exclude(producto__exact="")
        .values("producto")
        .annotate(total=Count("id"))
        .order_by("-total", "producto")[:10]
    )

    cuarteles_top = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total=Count("id"))
        .order_by("-total", "cuartel__nombre_cuartel")[:10]
    )

    monthly = build_month_series(
        queryset,
        date_field="fecha_aplicacion",
        aggregate_name="Aplicaciones",
        aggregate_expression=Count("id"),
        months=12,
    )

    return {
        "productos_top": {
            "labels": [item["producto"] for item in productos_top],
            "series": [item["total"] for item in productos_top],
        },
        "cuarteles_top": {
            "labels": [f"{item['cuartel__nombre_cuartel']} ({item['cuartel__predio__nombre_predio']})" for item in cuarteles_top],
            "series": [item["total"] for item in cuarteles_top],
        },
        "mensual": {
            "labels": monthly["labels"],
            "series": monthly["series"],
        },
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def fertilizacion_analytics_api(request):
    filters = get_dashboard_filters(request)
    qs = get_filtered_fertilizaciones(request)
    qs = apply_common_filters(qs, filters, "fecha_aplicacion")

    return JsonResponse(
        {
            "success": True,
            "resumen": build_fertilizacion_resumen(qs),
            "analytics": build_fertilizacion_analytics(qs),
        }
    )





@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def fertilizaciones_list_view(request):
    current_user = get_current_user(request)
    q = request.GET.get("q", "").strip()
    fertilizaciones = get_filtered_fertilizaciones(request, q)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")
    resumen = build_fertilizacion_resumen(fertilizaciones)

    return render(
        request,
        "fertilizaciones_lista.html",
        {
            "fertilizaciones": fertilizaciones,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "resumen": resumen,
            "q": q,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_fertilizaciones(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def fertilizacion_create_view(request):
    if request.method != "POST":
        return redirect("fertilizaciones_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_fertilizaciones(request),
        "No tienes permisos para crear fertilizaciones.",
    )
    if denied_response:
        return denied_response

    try:
        form = FertilizacionCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            fertilizacion = form.save()
            fertilizacion.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Aplicación de fertilización creada correctamente.",
                    "fertilizacion": serialize_fertilizacion(fertilizacion),
                    "resumen": build_fertilizacion_resumen(get_filtered_fertilizaciones(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def fertilizacion_detail_view(request, pk):
    if not user_owns_fertilizacion(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    try:
        fertilizacion = Fertilizacion.objects.select_related("cuartel", "cuartel__predio", "cuartel__predio__usuario").get(pk=pk)
        return JsonResponse({"success": True, "fertilizacion": serialize_fertilizacion(fertilizacion)})
    except Fertilizacion.DoesNotExist:
        return JsonResponse({"success": False, "error": "Fertilización no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def fertilizacion_update_view(request, pk):
    if not user_owns_fertilizacion(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return redirect("fertilizaciones_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_fertilizaciones(request),
        "No tienes permisos para editar fertilizaciones.",
    )
    if denied_response:
        return denied_response

    try:
        fertilizacion = Fertilizacion.objects.get(pk=pk)
        form = FertilizacionUpdateForm(normalize_estado_post_data(request), instance=fertilizacion, request=request)
        if form.is_valid():
            fertilizacion = form.save()
            fertilizacion.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Aplicación de fertilización actualizada correctamente.",
                    "fertilizacion": serialize_fertilizacion(fertilizacion),
                    "resumen": build_fertilizacion_resumen(get_filtered_fertilizaciones(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Fertilizacion.DoesNotExist:
        return JsonResponse({"success": False, "error": "Fertilización no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def fertilizacion_delete_view(request, pk):
    if not user_owns_fertilizacion(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_fertilizaciones(request),
        "No tienes permisos para eliminar fertilizaciones.",
    )
    if denied_response:
        return denied_response

    try:
        fertilizacion = Fertilizacion.objects.select_related("cuartel").get(pk=pk)
        nombre_cuartel = fertilizacion.cuartel.nombre_cuartel
        fertilizacion.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Aplicación de fertilización del cuartel {nombre_cuartel} eliminada correctamente.",
                "resumen": build_fertilizacion_resumen(get_filtered_fertilizaciones(request)),
            }
        )
    except Fertilizacion.DoesNotExist:
        return JsonResponse({"success": False, "error": "Fertilización no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def fertilizacion_toggle_estado_view(request, pk):
    if not user_owns_fertilizacion(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_fertilizaciones(request),
        "No tienes permisos para cambiar estado de fertilizaciones.",
    )
    if denied_response:
        return denied_response

    try:
        fertilizacion = Fertilizacion.objects.get(pk=pk)
        fertilizacion.estado = not bool(fertilizacion.estado)
        fertilizacion.save(update_fields=["estado"])
        estado_txt = "activada" if fertilizacion.estado else "desactivada"
        return JsonResponse(
            {
                "success": True,
                "message": f"Aplicación de fertilización {fertilizacion.id} {estado_txt} correctamente.",
                "estado": bool(fertilizacion.estado),
                "resumen": build_fertilizacion_resumen(get_filtered_fertilizaciones(request)),
            }
        )
    except Fertilizacion.DoesNotExist:
        return JsonResponse({"success": False, "error": "Fertilización no encontrada"}, status=404)


# ---------------------------------------------------------------------------
# COSECHAS
# ---------------------------------------------------------------------------

def build_cosecha_resumen(queryset=None):
    qs = queryset if queryset is not None else Cosecha.objects.all()
    total_cosechas = qs.count()
    total_kg = qs.aggregate(total=Sum("cantidad_kg"))["total"] or 0
    total_bins = qs.aggregate(total=Sum("cantidad_bins"))["total"] or 0
    ultima_cosecha_obj = qs.select_related("cuartel").order_by("-fecha_cosecha", "-created_at", "-id").first()
    ultima_cosecha = (
        f"{ultima_cosecha_obj.cuartel.nombre_cuartel} - {ultima_cosecha_obj.fecha_cosecha.strftime('%d/%m/%Y')}"
        if ultima_cosecha_obj
        else "-"
    )

    return {
        "total_cosechas": total_cosechas,
        "total_kg_cosechados": str(total_kg),
        "total_bins_cosechados": str(total_bins),
        "ultima_cosecha": ultima_cosecha,
    }


def serialize_cosecha(cosecha):
    return {
        "id": cosecha.id,
        "predio_id": cosecha.cuartel.predio_id,
        "predio_nombre": cosecha.cuartel.predio.nombre_predio,
        "cuartel_id": cosecha.cuartel_id,
        "cuartel_nombre": cosecha.cuartel.nombre_cuartel,
        "fecha_cosecha": cosecha.fecha_cosecha.strftime("%d/%m/%Y") if cosecha.fecha_cosecha else "-",
        "fecha_cosecha_iso": cosecha.fecha_cosecha.strftime("%Y-%m-%d") if cosecha.fecha_cosecha else "",
        "tipo_cosecha": cosecha.tipo_cosecha or "",
        "cantidad_kg": str(cosecha.cantidad_kg) if cosecha.cantidad_kg is not None else "",
        "cantidad_bins": str(cosecha.cantidad_bins) if cosecha.cantidad_bins is not None else "",
        "calidad": cosecha.calidad or "",
        "calidad_label": cosecha.get_calidad_display() if cosecha.calidad else "-",
        "destino": cosecha.destino or "",
        "destino_label": cosecha.get_destino_display() if cosecha.destino else "-",
        "observaciones": cosecha.observaciones or "",
        "estado": bool(cosecha.estado),
        "estado_label": "Activa" if cosecha.estado else "Inactiva",
        "created_at": cosecha.created_at.strftime("%d/%m/%Y %H:%M") if cosecha.created_at else "-",
        **build_responsable_info(cosecha.cuartel.predio.usuario),
        **build_ubicacion_info(cosecha.cuartel.predio, cosecha.cuartel.nombre_cuartel),
    }


_DECIMAL_ZERO = Value(Decimal("0.00"), output_field=DecimalField())

def build_cosecha_analytics(queryset):
    produccion_cuartel = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total_kg=Coalesce(Sum("cantidad_kg"), _DECIMAL_ZERO))
        .order_by("-total_kg", "cuartel__nombre_cuartel")[:10]
    )

    produccion_variedad = list(
        queryset.values("cuartel__variedad")
        .annotate(total_kg=Coalesce(Sum("cantidad_kg"), _DECIMAL_ZERO))
        .order_by("-total_kg", "cuartel__variedad")[:10]
    )

    rendimiento = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total_kg=Coalesce(Sum("cantidad_kg"), _DECIMAL_ZERO), total_bins=Coalesce(Sum("cantidad_bins"), _DECIMAL_ZERO))
        .order_by("-total_kg", "-total_bins")[:10]
    )

    monthly = build_month_series(
        queryset,
        date_field="fecha_cosecha",
        aggregate_name="KG",
        aggregate_expression=Coalesce(Sum("cantidad_kg"), _DECIMAL_ZERO),
        months=12,
    )

    return {
        "produccion_cuartel": {
            "labels": [f"{item['cuartel__nombre_cuartel']} ({item['cuartel__predio__nombre_predio']})" for item in produccion_cuartel],
            "series": [round(float(item["total_kg"] or 0), 2) for item in produccion_cuartel],
        },
        "produccion_variedad": {
            "labels": [item["cuartel__variedad"] or "Sin variedad" for item in produccion_variedad],
            "series": [round(float(item["total_kg"] or 0), 2) for item in produccion_variedad],
        },
        "produccion_mensual": {
            "labels": monthly["labels"],
            "series": monthly["series"],
        },
        "ranking_rendimiento": [
            {
                "cuartel": item["cuartel__nombre_cuartel"],
                "predio": item["cuartel__predio__nombre_predio"],
                "kg": round(float(item["total_kg"] or 0), 2),
                "bins": round(float(item["total_bins"] or 0), 2),
            }
            for item in rendimiento
        ],
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cosecha_analytics_api(request):
    filters = get_dashboard_filters(request)
    qs = get_filtered_cosechas(request)
    qs = apply_common_filters(qs, filters, "fecha_cosecha")

    return JsonResponse(
        {
            "success": True,
            "resumen": build_cosecha_resumen(qs),
            "analytics": build_cosecha_analytics(qs),
        }
    )





@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cosechas_list_view(request):
    current_user = get_current_user(request)
    q = request.GET.get("q", "").strip()
    cosechas = get_filtered_cosechas(request, q)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")
    resumen = build_cosecha_resumen(cosechas)

    return render(
        request,
        "cosechas_lista.html",
        {
            "cosechas": cosechas,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "resumen": resumen,
            "q": q,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_cosechas(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cosecha_create_view(request):
    if request.method != "POST":
        return redirect("cosechas_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_cosechas(request),
        "No tienes permisos para crear cosechas.",
    )
    if denied_response:
        return denied_response

    try:
        form = CosechaCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            cosecha = form.save()
            cosecha.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Cosecha creada correctamente.",
                    "cosecha": serialize_cosecha(cosecha),
                    "resumen": build_cosecha_resumen(get_filtered_cosechas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def cosecha_detail_view(request, pk):
    if not user_owns_cosecha(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    try:
        cosecha = Cosecha.objects.select_related("cuartel", "cuartel__predio", "cuartel__predio__usuario").get(pk=pk)
        return JsonResponse({"success": True, "cosecha": serialize_cosecha(cosecha)})
    except Cosecha.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cosecha no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cosecha_update_view(request, pk):
    if not user_owns_cosecha(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return redirect("cosechas_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_cosechas(request),
        "No tienes permisos para editar cosechas.",
    )
    if denied_response:
        return denied_response

    try:
        cosecha = Cosecha.objects.get(pk=pk)
        form = CosechaUpdateForm(normalize_estado_post_data(request), instance=cosecha, request=request)
        if form.is_valid():
            cosecha = form.save()
            cosecha.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Cosecha actualizada correctamente.",
                    "cosecha": serialize_cosecha(cosecha),
                    "resumen": build_cosecha_resumen(get_filtered_cosechas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Cosecha.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cosecha no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cosecha_delete_view(request, pk):
    if not user_owns_cosecha(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_cosechas(request),
        "No tienes permisos para eliminar cosechas.",
    )
    if denied_response:
        return denied_response

    try:
        cosecha = Cosecha.objects.select_related("cuartel").get(pk=pk)
        nombre_cuartel = cosecha.cuartel.nombre_cuartel
        cosecha.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Cosecha del cuartel {nombre_cuartel} eliminada correctamente.",
                "resumen": build_cosecha_resumen(get_filtered_cosechas(request)),
            }
        )
    except Cosecha.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cosecha no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def cosecha_toggle_estado_view(request, pk):
    if not user_owns_cosecha(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_cosechas(request),
        "No tienes permisos para cambiar estado de cosechas.",
    )
    if denied_response:
        return denied_response

    try:
        cosecha = Cosecha.objects.get(pk=pk)
        cosecha.estado = not bool(cosecha.estado)
        cosecha.save(update_fields=["estado"])
        estado_txt = "activada" if cosecha.estado else "desactivada"
        return JsonResponse(
            {
                "success": True,
                "message": f"Cosecha {cosecha.id} {estado_txt} correctamente.",
                "estado": bool(cosecha.estado),
                "resumen": build_cosecha_resumen(get_filtered_cosechas(request)),
            }
        )
    except Cosecha.DoesNotExist:
        return JsonResponse({"success": False, "error": "Cosecha no encontrada"}, status=404)


# ---------------------------------------------------------------------------
# APLICACIONES QUIMICAS
# ---------------------------------------------------------------------------

def build_aplicacion_quimica_resumen(queryset=None):
    qs = queryset if queryset is not None else AplicacionQuimica.objects.all()
    total_aplicaciones = qs.count()
    productos_utilizados = qs.exclude(producto__isnull=True).exclude(producto__exact="").values("producto").distinct().count()
    cuarteles_tratados = qs.values("cuartel_id").distinct().count()
    ultima_aplicacion_obj = qs.select_related("cuartel").order_by("-fecha_aplicacion", "-created_at", "-id").first()
    ultima_aplicacion = (
        f"{ultima_aplicacion_obj.cuartel.nombre_cuartel} - {ultima_aplicacion_obj.fecha_aplicacion.strftime('%d/%m/%Y')}"
        if ultima_aplicacion_obj
        else "-"
    )

    return {
        "total_aplicaciones": total_aplicaciones,
        "productos_utilizados": productos_utilizados,
        "cuarteles_tratados": cuarteles_tratados,
        "ultima_aplicacion": ultima_aplicacion,
    }


def serialize_aplicacion_quimica(aplicacion):
    return {
        "id": aplicacion.id,
        "predio_id": aplicacion.cuartel.predio_id,
        "predio_nombre": aplicacion.cuartel.predio.nombre_predio,
        "cuartel_id": aplicacion.cuartel_id,
        "cuartel_nombre": aplicacion.cuartel.nombre_cuartel,
        "fecha_aplicacion": aplicacion.fecha_aplicacion.strftime("%d/%m/%Y") if aplicacion.fecha_aplicacion else "-",
        "fecha_aplicacion_iso": aplicacion.fecha_aplicacion.strftime("%Y-%m-%d") if aplicacion.fecha_aplicacion else "",
        "producto": aplicacion.producto or "",
        "tipo_producto": aplicacion.tipo_producto or "",
        "dosis": str(aplicacion.dosis) if aplicacion.dosis is not None else "",
        "unidad": aplicacion.unidad or "",
        "metodo_aplicacion": aplicacion.metodo_aplicacion or "",
        "responsable": aplicacion.responsable or "",
        "observaciones": aplicacion.observaciones or "",
        "estado": bool(aplicacion.estado),
        "created_at": aplicacion.created_at.strftime("%d/%m/%Y %H:%M") if aplicacion.created_at else "-",
        **build_responsable_info(aplicacion.cuartel.predio.usuario),
        **build_ubicacion_info(aplicacion.cuartel.predio, aplicacion.cuartel.nombre_cuartel),
    }


def build_aplicacion_quimica_analytics(queryset):
    productos_top = list(
        queryset.exclude(producto__isnull=True)
        .exclude(producto__exact="")
        .values("producto")
        .annotate(total=Count("id"))
        .order_by("-total", "producto")[:10]
    )

    cuarteles_top = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total=Count("id"))
        .order_by("-total", "cuartel__nombre_cuartel")[:10]
    )

    monthly = build_month_series(
        queryset,
        date_field="fecha_aplicacion",
        aggregate_name="Aplicaciones",
        aggregate_expression=Count("id"),
        months=12,
    )

    return {
        "productos_top": {
            "labels": [item["producto"] for item in productos_top],
            "series": [item["total"] for item in productos_top],
        },
        "cuarteles_top": {
            "labels": [f"{item['cuartel__nombre_cuartel']} ({item['cuartel__predio__nombre_predio']})" for item in cuarteles_top],
            "series": [item["total"] for item in cuarteles_top],
        },
        "mensual": {
            "labels": monthly["labels"],
            "series": monthly["series"],
        },
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_analytics_api(request):
    filters = get_dashboard_filters(request)
    qs = get_filtered_aplicaciones_quimicas(request)
    qs = apply_common_filters(qs, filters, "fecha_aplicacion")

    return JsonResponse(
        {
            "success": True,
            "resumen": build_aplicacion_quimica_resumen(qs),
            "analytics": build_aplicacion_quimica_analytics(qs),
        }
    )





@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def aplicaciones_quimicas_list_view(request):
    current_user = get_current_user(request)
    q = request.GET.get("q", "").strip()
    aplicaciones = get_filtered_aplicaciones_quimicas(request, q)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")
    resumen = build_aplicacion_quimica_resumen(aplicaciones)

    return render(
        request,
        "aplicaciones_quimicas_lista.html",
        {
            "aplicaciones": aplicaciones,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "resumen": resumen,
            "q": q,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_aplicaciones_quimicas(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_create_view(request):
    if request.method != "POST":
        return redirect("aplicaciones_quimicas_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_aplicaciones_quimicas(request),
        "No tienes permisos para crear aplicaciones quimicas.",
    )
    if denied_response:
        return denied_response

    try:
        form = AplicacionQuimicaCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            aplicacion = form.save()
            aplicacion.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Aplicación química creada correctamente.",
                    "aplicacion": serialize_aplicacion_quimica(aplicacion),
                    "resumen": build_aplicacion_quimica_resumen(get_filtered_aplicaciones_quimicas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_detail_view(request, pk):
    if not user_owns_aplicacion_quimica(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    try:
        aplicacion = AplicacionQuimica.objects.select_related("cuartel", "cuartel__predio", "cuartel__predio__usuario").get(pk=pk)
        return JsonResponse({"success": True, "aplicacion": serialize_aplicacion_quimica(aplicacion)})
    except AplicacionQuimica.DoesNotExist:
        return JsonResponse({"success": False, "error": "Aplicación química no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_update_view(request, pk):
    if not user_owns_aplicacion_quimica(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return redirect("aplicaciones_quimicas_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_aplicaciones_quimicas(request),
        "No tienes permisos para editar aplicaciones quimicas.",
    )
    if denied_response:
        return denied_response

    try:
        aplicacion = AplicacionQuimica.objects.get(pk=pk)
        form = AplicacionQuimicaUpdateForm(normalize_estado_post_data(request), instance=aplicacion, request=request)
        if form.is_valid():
            aplicacion = form.save()
            aplicacion.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Aplicación química actualizada correctamente.",
                    "aplicacion": serialize_aplicacion_quimica(aplicacion),
                    "resumen": build_aplicacion_quimica_resumen(get_filtered_aplicaciones_quimicas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except AplicacionQuimica.DoesNotExist:
        return JsonResponse({"success": False, "error": "Aplicación química no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_delete_view(request, pk):
    if not user_owns_aplicacion_quimica(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_aplicaciones_quimicas(request),
        "No tienes permisos para eliminar aplicaciones quimicas.",
    )
    if denied_response:
        return denied_response

    try:
        aplicacion = AplicacionQuimica.objects.select_related("cuartel").get(pk=pk)
        nombre_cuartel = aplicacion.cuartel.nombre_cuartel
        aplicacion.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Aplicación química del cuartel {nombre_cuartel} eliminada correctamente.",
                "resumen": build_aplicacion_quimica_resumen(get_filtered_aplicaciones_quimicas(request)),
            }
        )
    except AplicacionQuimica.DoesNotExist:
        return JsonResponse({"success": False, "error": "Aplicación química no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def aplicacion_quimica_toggle_estado_view(request, pk):
    if not user_owns_aplicacion_quimica(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
    
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_aplicaciones_quimicas(request),
        "No tienes permisos para cambiar estado de aplicaciones quimicas.",
    )
    if denied_response:
        return denied_response

    try:
        aplicacion = AplicacionQuimica.objects.get(pk=pk)
        aplicacion.estado = not bool(aplicacion.estado)
        aplicacion.save(update_fields=["estado"])
        estado_txt = "activada" if aplicacion.estado else "desactivada"
        return JsonResponse(
            {
                "success": True,
                "message": f"Aplicación química {aplicacion.id} {estado_txt} correctamente.",
                "estado": bool(aplicacion.estado),
                "resumen": build_aplicacion_quimica_resumen(get_filtered_aplicaciones_quimicas(request)),
            }
        )
    except AplicacionQuimica.DoesNotExist:
        return JsonResponse({"success": False, "error": "Aplicación química no encontrada"}, status=404)


# ---------------------------------------------------------------------------
# LABORES AGRICOLAS
# ---------------------------------------------------------------------------

def serialize_labor_agricola(labor):
    return {
        "id": labor.id,
        "usuario_id": labor.usuario_id,
        "usuario_nombre": labor.usuario.nombre if labor.usuario_id else "-",
        "predio_id": labor.predio_id,
        "predio_nombre": labor.predio.nombre_predio,
        "cuartel_id": labor.cuartel_id,
        "cuartel_nombre": labor.cuartel.nombre_cuartel,
        "fecha": labor.fecha.strftime("%d/%m/%Y") if labor.fecha else "-",
        "fecha_iso": labor.fecha.strftime("%Y-%m-%d") if labor.fecha else "",
        "tipo_labor": labor.tipo_labor or "",
        "tipo_labor_label": labor.get_tipo_labor_display() if labor.tipo_labor else "",
        "subtipo": labor.subtipo or "",
        "responsable": labor.responsable or "",
        "descripcion": labor.descripcion or "",
        "observaciones": labor.observaciones or "",
        "estado": bool(labor.estado),
        "created_at": labor.created_at.strftime("%d/%m/%Y %H:%M") if labor.created_at else "-",
        **build_responsable_info(labor.usuario if labor.usuario_id else labor.predio.usuario),
        **build_ubicacion_info(labor.predio, labor.cuartel.nombre_cuartel if labor.cuartel_id else ""),
    }


def build_labores_resumen(queryset=None):
    qs = queryset if queryset is not None else LaborAgricola.objects.all()
    total_labores = qs.count()
    cuarteles_intervenidos = qs.values("cuartel_id").distinct().count()
    responsables = qs.exclude(responsable__isnull=True).exclude(responsable__exact="").values("responsable").distinct().count()
    ultima_labor_obj = qs.select_related("cuartel").order_by("-fecha", "-id").first()
    ultima_labor = (
        f"{ultima_labor_obj.cuartel.nombre_cuartel} - {ultima_labor_obj.fecha.strftime('%d/%m/%Y')}"
        if ultima_labor_obj
        else "-"
    )
    return {
        "total_labores": total_labores,
        "cuarteles_intervenidos": cuarteles_intervenidos,
        "responsables_activos": responsables,
        "ultima_labor": ultima_labor,
    }


def build_labores_analytics(queryset):
    frecuencia_tipos = list(
        queryset.values("tipo_labor")
        .annotate(total=Count("id"))
        .order_by("-total", "tipo_labor")
    )
    por_cuartel = list(
        queryset.values("cuartel__nombre_cuartel", "cuartel__predio__nombre_predio")
        .annotate(total=Count("id"))
        .order_by("-total", "cuartel__nombre_cuartel")[:10]
    )
    por_temporada = list(
        queryset.annotate(periodo=TruncMonth("fecha"))
        .values("periodo")
        .annotate(total=Count("id"))
        .order_by("periodo")
    )

    return {
        "frecuencia_labores": {
            "labels": [item["tipo_labor"].replace("_", " ").title() for item in frecuencia_tipos],
            "series": [item["total"] for item in frecuencia_tipos],
        },
        "labores_por_cuartel": {
            "labels": [f"{item['cuartel__nombre_cuartel']} ({item['cuartel__predio__nombre_predio']})" for item in por_cuartel],
            "series": [item["total"] for item in por_cuartel],
        },
        "labores_por_temporada": {
            "labels": [item["periodo"].strftime("%m/%Y") if item["periodo"] else "-" for item in por_temporada],
            "series": [item["total"] for item in por_temporada],
        },
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labores_agricolas_list_view(request):
    current_user = get_current_user(request)
    q = request.GET.get("q", "").strip()
    labores = get_filtered_labores_agricolas(request, q)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")

    return render(
        request,
        "labores_agricolas_lista.html",
        {
            "labores": labores,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "resumen": build_labores_resumen(labores),
            "q": q,
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_labores_agricolas(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labores_agricolas_analitica_view(request):
    current_user = get_current_user(request)
    labores = get_filtered_labores_agricolas(request)
    predios_activos = get_filtered_predios(request).filter(estado=True).order_by("nombre_predio")
    cuarteles_activos = get_filtered_cuarteles(request).filter(estado=True, predio__estado=True).order_by("nombre_cuartel")

    return render(
        request,
        "labores_agricolas_lista.html",
        {
            "labores": labores,
            "predios_activos": predios_activos,
            "cuarteles_activos": cuarteles_activos,
            "current_user": current_user,
            "analytics": build_labores_analytics(labores),
            "q": "",
            "sidebar_context": get_sidebar_context(request),
            "can_manage": can_manage_labores_agricolas(request),
        },
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labores_agricolas_analytics_api(request):
    filters = get_dashboard_filters(request)
    qs = get_filtered_labores_agricolas(request)

    predio_id = filters.get("predio_id")
    cuartel_id = filters.get("cuartel_id")
    fecha_desde = filters.get("fecha_desde")
    fecha_hasta = filters.get("fecha_hasta")

    if predio_id:
        qs = qs.filter(predio_id=predio_id)
    if cuartel_id:
        qs = qs.filter(cuartel_id=cuartel_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    return JsonResponse(
        {
            "success": True,
            "resumen": build_labores_resumen(qs),
            "analytics": build_labores_analytics(qs),
        }
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def labor_agricola_create_view(request):
    if request.method != "POST":
        return redirect("labores_agricolas_lista")

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_labores_agricolas(request),
        "No tienes permisos para crear labores agrícolas.",
    )
    if denied_response:
        return denied_response

    try:
        form = LaborAgricolaCreateForm(normalize_estado_post_data(request), request=request)
        if form.is_valid():
            labor = form.save(commit=False)
            labor.usuario_id = request.session.get("usuario_id")
            labor.save()
            labor.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Labor agrícola creada correctamente.",
                    "labor": serialize_labor_agricola(labor),
                    "resumen": build_labores_resumen(get_filtered_labores_agricolas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labor_agricola_detail_view(request, pk):
    if not user_owns_labor_agricola(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)

    try:
        labor = LaborAgricola.objects.select_related("usuario", "predio", "predio__usuario", "cuartel").get(pk=pk)
        return JsonResponse({"success": True, "labor": serialize_labor_agricola(labor)})
    except LaborAgricola.DoesNotExist:
        return JsonResponse({"success": False, "error": "Labor agrícola no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def labor_agricola_update_view(request, pk):
    if request.method != "POST":
        return redirect("labores_agricolas_lista")

    if not user_owns_labor_agricola(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_labores_agricolas(request),
        "No tienes permisos para editar labores agrícolas.",
    )
    if denied_response:
        return denied_response

    try:
        labor = LaborAgricola.objects.get(pk=pk)
        form = LaborAgricolaUpdateForm(normalize_estado_post_data(request), instance=labor, request=request)
        if form.is_valid():
            labor = form.save()
            labor.refresh_from_db()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Labor agrícola actualizada correctamente.",
                    "labor": serialize_labor_agricola(labor),
                    "resumen": build_labores_resumen(get_filtered_labores_agricolas(request)),
                }
            )

        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})
    except LaborAgricola.DoesNotExist:
        return JsonResponse({"success": False, "error": "Labor agrícola no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def labor_agricola_delete_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_labor_agricola(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)

    denied_response = deny_if_cannot_manage(
        request,
        can_manage_labores_agricolas(request),
        "No tienes permisos para eliminar labores agrícolas.",
    )
    if denied_response:
        return denied_response

    try:
        labor = LaborAgricola.objects.select_related("cuartel").get(pk=pk)
        nombre_cuartel = labor.cuartel.nombre_cuartel
        labor.delete()
        return JsonResponse(
            {
                "success": True,
                "message": f"Labor agrícola del cuartel {nombre_cuartel} eliminada correctamente.",
                "resumen": build_labores_resumen(get_filtered_labores_agricolas(request)),
            }
        )
    except LaborAgricola.DoesNotExist:
        return JsonResponse({"success": False, "error": "Labor agrícola no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_PRODUCTOR)
def labor_agricola_toggle_estado_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not user_owns_labor_agricola(request, pk):
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)

    try:
        labor = LaborAgricola.objects.get(pk=pk)
        labor.estado = not bool(labor.estado)
        labor.save(update_fields=["estado"])
        estado_txt = "activada" if labor.estado else "desactivada"
        return JsonResponse(
            {
                "success": True,
                "message": f"Labor agrícola {labor.id} {estado_txt} correctamente.",
                "estado": bool(labor.estado),
                "resumen": build_labores_resumen(get_filtered_labores_agricolas(request)),
            }
        )
    except LaborAgricola.DoesNotExist:
        return JsonResponse({"success": False, "error": "Labor agrícola no encontrada"}, status=404)


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labores_agricolas_export_excel_view(request):
    if find_spec("openpyxl") is None:
        return build_export_dependency_error_response(request, ["openpyxl"])

    from openpyxl import Workbook

    labores = get_filtered_labores_agricolas(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Labores"
    ws.append(["Fecha", "Predio", "Cuartel", "Tipo", "Subtipo", "Responsable", "Descripción", "Observaciones", "Estado"])

    for labor in labores:
        ws.append([
            labor.fecha.strftime("%d/%m/%Y") if labor.fecha else "-",
            labor.predio.nombre_predio,
            labor.cuartel.nombre_cuartel,
            labor.get_tipo_labor_display(),
            labor.subtipo or "-",
            labor.responsable or "-",
            labor.descripcion or "-",
            labor.observaciones or "-",
            "Activo" if labor.estado else "Inactivo",
        ])

    filename = f"labores_agricolas_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def labores_agricolas_export_pdf_view(request):
    if find_spec("reportlab") is None:
        return build_export_dependency_error_response(request, ["reportlab"])

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    labores = get_filtered_labores_agricolas(request)
    filename = f"labores_agricolas_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.pdf"

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, y, "Reporte de Labores Agricolas")
    y -= 24
    c.setFont("Helvetica", 8)

    for labor in labores:
        line = (
            f"{labor.fecha.strftime('%d/%m/%Y')} | {labor.predio.nombre_predio} | {labor.cuartel.nombre_cuartel} | "
            f"{labor.get_tipo_labor_display()} | {labor.subtipo or '-'} | {labor.responsable or '-'}"
        )
        c.drawString(30, y, line[:150])
        y -= 14
        if y < 40:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 8)

    c.save()
    return response


# ---------------------------------------------------------------------------
# COMENTARIOS TECNICOS (PRODESAL)
# ---------------------------------------------------------------------------

MODULO_DETAIL_ROUTE = {
    "predio": "predios_detalle",
    "cuartel": "cuarteles_detalle",
    "riego": "riegos_detalle",
    "fertilizacion": "fertilizaciones_detalle",
    "cosecha": "cosechas_detalle",
    "aplicacion_quimica": "aplicaciones_quimicas_detalle",
    "labor_agricola": "labores_agricolas_detalle",
}


def serialize_comentario(comentario):
    return {
        "id": comentario.id,
        "usuario_prodesal": comentario.usuario_prodesal.nombre,
        "modulo": comentario.modulo,
        "modulo_label": comentario.get_modulo_display(),
        "objeto_id": comentario.objeto_id,
        "comentario": comentario.comentario,
        "leido": comentario.leido,
        "fecha": comentario.fecha.strftime("%d/%m/%Y %H:%M") if comentario.fecha else "-",
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO, Usuario.ROL_PRODUCTOR)
def comentarios_por_registro_view(request, modulo, objeto_id):
    if modulo not in MODULO_DETAIL_ROUTE:
        return JsonResponse({"success": False, "error": "Módulo no válido"}, status=400)

    if not user_can_view_registro(request, modulo, objeto_id):
        return access_denied_response(request, "No autorizado para consultar este registro.")

    comentarios = ComentarioTecnico.objects.select_related("usuario_prodesal").filter(
        modulo=modulo, objeto_id=objeto_id
    )

    current_user = get_current_user(request)
    if current_user and current_user.rol == Usuario.ROL_PRODUCTOR:
        ComentarioTecnico.objects.filter(modulo=modulo, objeto_id=objeto_id, productor=current_user, leido=False).update(
            leido=True
        )

    return JsonResponse(
        {
            "success": True,
            "comentarios": [serialize_comentario(c) for c in comentarios],
        }
    )


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO)
def comentario_create_view(request, modulo, objeto_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not can_comment(request):
        return access_denied_response(request, "No tienes permisos para dejar observaciones.")

    if modulo not in MODULO_DETAIL_ROUTE:
        return JsonResponse({"success": False, "error": "Módulo no válido"}, status=400)

    registro, productor_id = get_registro_y_productor(modulo, objeto_id)
    if registro is None or not productor_id:
        return JsonResponse({"success": False, "error": "Registro no encontrado"}, status=404)

    current_user = get_current_user(request)
    form = ComentarioTecnicoForm(request.POST)
    if form.is_valid():
        comentario = form.save(commit=False)
        comentario.usuario_prodesal = current_user
        comentario.productor_id = productor_id
        comentario.modulo = modulo
        comentario.objeto_id = objeto_id
        comentario.save()
        return JsonResponse(
            {
                "success": True,
                "message": "Observación registrada correctamente.",
                "comentario": serialize_comentario(comentario),
            }
        )

    errors = {field: error[0] for field, error in form.errors.items()}
    return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})


# ---------------------------------------------------------------------------
# NOTIFICACIONES
# ---------------------------------------------------------------------------

def serialize_notificacion(notificacion):
    return {
        "id": notificacion.id,
        "titulo": notificacion.titulo,
        "mensaje": notificacion.mensaje,
        "usuario_generador": notificacion.usuario_generador.nombre,
        "usuario_generador_rol": ROL_LABELS.get(notificacion.usuario_generador.rol, notificacion.usuario_generador.rol),
        "productor": notificacion.productor.nombre,
        "modulo": notificacion.modulo or "",
        "modulo_label": notificacion.get_modulo_display() if notificacion.modulo else "",
        "objeto_id": notificacion.objeto_id,
        "leido": notificacion.leido,
        "fecha": notificacion.fecha.strftime("%d/%m/%Y") if notificacion.fecha else "-",
        "hora": notificacion.fecha.strftime("%H:%M") if notificacion.fecha else "-",
        "fecha_hora": notificacion.fecha.strftime("%d/%m/%Y %H:%M") if notificacion.fecha else "-",
    }


@role_required(Usuario.ROL_ADMIN, Usuario.ROL_TECNICO)
def notificacion_create_view(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    if not can_send_notifications(request):
        return access_denied_response(request, "No tienes permisos para enviar notificaciones.")

    modulo = request.POST.get("modulo", "").strip()
    objeto_id = request.POST.get("objeto_id", "").strip()
    productor = None

    if modulo and objeto_id:
        registro, productor_id_resuelto = get_registro_y_productor(modulo, objeto_id)
        if registro is None or not productor_id_resuelto:
            return JsonResponse({"success": False, "error": "Registro no encontrado"}, status=404)
        productor = Usuario.objects.filter(id=productor_id_resuelto, rol=Usuario.ROL_PRODUCTOR, estado=True).first()
    else:
        modulo = None
        objeto_id = None
        productor_id = request.POST.get("productor_id")
        productor = Usuario.objects.filter(id=productor_id, rol=Usuario.ROL_PRODUCTOR, estado=True).first()

    if not productor:
        return JsonResponse({"success": False, "error": "Productor no válido"}, status=400)

    current_user = get_current_user(request)
    form = NotificacionForm(request.POST)
    if form.is_valid():
        notificacion = form.save(commit=False)
        notificacion.usuario_generador = current_user
        notificacion.productor = productor
        notificacion.modulo = modulo
        notificacion.objeto_id = objeto_id or None
        notificacion.save()
        return JsonResponse(
            {
                "success": True,
                "message": "Notificación enviada correctamente.",
                "notificacion": serialize_notificacion(notificacion),
            }
        )

    errors = {field: error[0] for field, error in form.errors.items()}
    return JsonResponse({"success": False, "error": "Validación fallida", "errors": errors})


@login_required_custom
def notificaciones_list_view(request):
    current_user = get_current_user(request)
    if not current_user:
        return JsonResponse({"success": False, "error": "Sesión inválida"}, status=401)

    if current_user.rol == Usuario.ROL_PRODUCTOR:
        notificaciones = Notificacion.objects.select_related("usuario_generador").filter(productor=current_user)
    elif current_user.rol == Usuario.ROL_TECNICO:
        notificaciones = Notificacion.objects.select_related("usuario_generador", "productor").filter(
            usuario_generador=current_user
        )
    else:
        notificaciones = Notificacion.objects.select_related("usuario_generador", "productor").all()

    return JsonResponse(
        {
            "success": True,
            "notificaciones": [serialize_notificacion(n) for n in notificaciones[:50]],
            "no_leidas": notificaciones.filter(leido=False).count() if current_user.rol == Usuario.ROL_PRODUCTOR else 0,
        }
    )


@login_required_custom
def notificacion_marcar_leida_view(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    current_user = get_current_user(request)
    if not current_user:
        return JsonResponse({"success": False, "error": "Sesión inválida"}, status=401)

    try:
        notificacion = Notificacion.objects.get(pk=pk, productor=current_user)
    except Notificacion.DoesNotExist:
        return JsonResponse({"success": False, "error": "Notificación no encontrada"}, status=404)

    notificacion.leido = True
    notificacion.save(update_fields=["leido"])
    return JsonResponse({"success": True})

